# app/modules/reports/excel.py
# ================================================================
# Auto-D Kenya - Excel Report Generator
# ================================================================

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment,
)
from openpyxl.utils import get_column_letter


class ExcelGenerator:
    """Professional Excel Report Generator."""

    DARK = "0F172A"
    GOLD = "F59E0B"
    GREEN = "16A34A"
    LIGHT = "F8FAFC"
    BORDER = "CBD5E1"

    @staticmethod
    def money(value):
        try:
            return float(value)
        except Exception:
            return 0

    @staticmethod
    def generate_valuation_report(report: dict) -> bytes:

        wb = Workbook()
        ws = wb.active
        ws.title = "Valuation Summary"

        valuation = report.get("valuation", {})
        vehicle = valuation.get("vehicle", {})
        explanation = valuation.get("price_explanation", {})
        value_range = valuation.get("estimated_value_range", {})

        # ----------------------------------------------------
        # Styles
        # ----------------------------------------------------

        title_font = Font(size=20, bold=True, color=ExcelGenerator.DARK)
        heading_font = Font(size=13, bold=True, color="FFFFFF")
        label_font = Font(bold=True)
        value_font = Font(size=24, bold=True, color=ExcelGenerator.GREEN)

        gold_fill = PatternFill(
            fill_type="solid",
            start_color=ExcelGenerator.GOLD,
            end_color=ExcelGenerator.GOLD,
        )

        light_fill = PatternFill(
            fill_type="solid",
            start_color=ExcelGenerator.LIGHT,
            end_color=ExcelGenerator.LIGHT,
        )

        thin = Side(style="thin", color=ExcelGenerator.BORDER)

        border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin,
        )

        center = Alignment(horizontal="center")
        left = Alignment(horizontal="left")

        row = 1

        # ----------------------------------------------------
        # Header
        # ----------------------------------------------------

        ws.merge_cells("A1:F1")
        ws["A1"] = "AUTO-D KENYA"
        ws["A1"].font = title_font
        ws["A1"].alignment = center

        row += 1

        ws.merge_cells("A2:F2")
        ws["A2"] = "Vehicle Valuation Report"
        ws["A2"].alignment = center
        ws["A2"].font = Font(size=16, bold=True)

        row = 4

        ws["A4"] = "Generated"
        ws["B4"] = datetime.now().strftime("%d %B %Y %H:%M")

        row = 6

        # ----------------------------------------------------
        # Estimated Value
        # ----------------------------------------------------

        ws.merge_cells("A6:F6")
        ws["A6"] = "ESTIMATED MARKET VALUE"
        ws["A6"].fill = gold_fill
        ws["A6"].font = heading_font
        ws["A6"].alignment = center

        row = 7

        ws.merge_cells("A7:F7")

        value = valuation.get(
            "estimated_vehicle_value",
            valuation.get("market_value", 0),
        )

        ws["A7"] = value
        ws["A7"].number_format = '"KES" #,##0'
        ws["A7"].font = value_font
        ws["A7"].alignment = center

        row = 9

        ws["A9"] = "Confidence"
        ws["B9"] = f"{valuation.get('confidence_score',0)}%"

        ws["D9"] = "Value Range"

        ws["E9"] = (
            f"KES {value_range.get('minimum',0):,.0f}"
            f" - "
            f"KES {value_range.get('maximum',0):,.0f}"
        )

        row = 11

        # ----------------------------------------------------
        # Vehicle Profile
        # ----------------------------------------------------

        ws["A11"] = "Vehicle Profile"
        ws["A11"].fill = gold_fill
        ws["A11"].font = heading_font

        profile = [
            ("Make", vehicle.get("make")),
            ("Model", vehicle.get("model")),
            ("Variant", vehicle.get("variant")),
            ("Year", vehicle.get("year")),
            ("Mileage", f"{vehicle.get('mileage',0):,} km"),
            ("Fuel", vehicle.get("fuel_type")),
            ("Transmission", vehicle.get("transmission")),
            ("Engine", f"{vehicle.get('engine_size_cc',0)} cc"),
            ("Body", vehicle.get("body_type")),
            ("Condition", vehicle.get("condition")),
            ("Location", vehicle.get("location")),
        ]

        row = 12

        for label, value in profile:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value

            ws[f"A{row}"].font = label_font
            ws[f"A{row}"].fill = light_fill

            ws[f"A{row}"].border = border
            ws[f"B{row}"].border = border

            row += 1

        row += 1

        # ----------------------------------------------------
        # Values
        # ----------------------------------------------------

        ws[f"A{row}"] = "Valuation Results"
        ws[f"A{row}"].fill = gold_fill
        ws[f"A{row}"].font = heading_font

        row += 1

        values = [
            ("Market Value", valuation.get("market_value")),
            ("Retail Value", valuation.get("retail_value")),
            ("Dealer Value", valuation.get("dealer_value")),
            ("Trade Value", valuation.get("trade_value")),
            ("Base Price", valuation.get("base_price")),
        ]

        for label, value in values:

            ws[f"A{row}"] = label
            ws[f"B{row}"] = value

            ws[f"A{row}"].font = label_font
            ws[f"A{row}"].fill = light_fill

            ws[f"B{row}"].number_format = '"KES" #,##0'

            ws[f"A{row}"].border = border
            ws[f"B{row}"].border = border

            row += 1

        row += 1

        # ----------------------------------------------------
        # Factors
        # ----------------------------------------------------

        ws[f"A{row}"] = "Adjustment Factors"
        ws[f"A{row}"].fill = gold_fill
        ws[f"A{row}"].font = heading_font

        row += 1

        factors = [
            ("Mileage", valuation.get("mileage_factor")),
            ("Condition", valuation.get("condition_factor")),
            ("Accident", valuation.get("accident_factor")),
            ("Location", valuation.get("location_factor")),
            ("Demand", valuation.get("demand_factor")),
            ("Trend", valuation.get("trend_factor")),
            ("Features", valuation.get("feature_factor")),
        ]

        for label, value in factors:

            ws[f"A{row}"] = label
            ws[f"B{row}"] = value

            ws[f"A{row}"].font = label_font
            ws[f"A{row}"].fill = light_fill

            ws[f"A{row}"].border = border
            ws[f"B{row}"].border = border

            row += 1

        row += 1

        # ----------------------------------------------------
        # Market Information
        # ----------------------------------------------------

        ws[f"A{row}"] = "Market Information"
        ws[f"A{row}"].fill = gold_fill
        ws[f"A{row}"].font = heading_font

        row += 1

        market = [
            ("Base Price Source", valuation.get("base_price_source")),
            ("Listing Count", valuation.get("listing_count")),
            ("Confidence", valuation.get("confidence_score")),
        ]

        for label, value in market:

            ws[f"A{row}"] = label
            ws[f"B{row}"] = value

            ws[f"A{row}"].font = label_font
            ws[f"A{row}"].fill = light_fill

            ws[f"A{row}"].border = border
            ws[f"B{row}"].border = border

            row += 1

        row += 1

        # ----------------------------------------------------
        # Summary
        # ----------------------------------------------------

        ws[f"A{row}"] = "Valuation Summary"
        ws[f"A{row}"].fill = gold_fill
        ws[f"A{row}"].font = heading_font

        row += 1

        ws.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row + 3,
            end_column=6,
        )

        ws.cell(row=row, column=1).value = explanation.get(
            "summary",
            "Vehicle valued using AUTO-D valuation engine."
        )

        ws.cell(row=row, column=1).alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )

        row += 5

        # ----------------------------------------------------
        # Disclaimer
        # ----------------------------------------------------

        ws[f"A{row}"] = "Disclaimer"
        ws[f"A{row}"].fill = gold_fill
        ws[f"A{row}"].font = heading_font

        row += 1

        ws.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row + 4,
            end_column=6,
        )

        ws.cell(row=row, column=1).value = (
            "This valuation represents an indicative market value "
            "generated by the AUTO-D Kenya valuation engine using "
            "market data, depreciation models, mileage, condition "
            "and regional demand. Actual transaction values may vary."
        )

        ws.cell(row=row, column=1).alignment = Alignment(
            wrap_text=True
        )

        # ----------------------------------------------------
        # Formatting
        # ----------------------------------------------------

        ws.freeze_panes = "A6"

        for col in range(1, 7):

            width = 20

            if col == 2:
                width = 35

            ws.column_dimensions[
                get_column_letter(col)
            ].width = width

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer.getvalue()
